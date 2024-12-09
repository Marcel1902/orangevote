from django.contrib import admin
from django.urls import path
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from vote.models import Commune, Vote, Zone, SousZone, Image
from vote.views import commune
  # Assurez-vous que ces modèles sont bien importés

# Fonction d'exportation des résultats
def export_vote_results(modeladmin, request, queryset):
    # Vérification des permissions de l'utilisateur
    if not request.user.is_superuser:
        return HttpResponse("Vous n'avez pas les permissions nécessaires pour effectuer cette action.")

    # Créer un nouveau classeur Excel
    wb = Workbook()

    # Fonction pour appliquer un format d'en-tête
    def apply_header_format(ws):
        for cell in ws[1]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)

    # Feuille des résultats des communes
    ws_commune = wb.active
    ws_commune.title = "Résultats des Communes"
    headers_commune = ['Commune', 'Qualité du Site', 'Originalité du Support', 'site_brandes', 'repris_concurrence',
                       'rapidite_deploiement', 'Total']
    ws_commune.append(headers_commune)
    apply_header_format(ws_commune)

    # Résultats des communes regroupés
    commune_results = {}

    # Boucle pour récupérer et agréger les résultats des communes
    for commune in queryset:
        if commune.name not in commune_results:
            commune_results[commune.name] = {
                'qualite_site': 0,
                'originalite_support': 0,
                'site_brandes': commune.site_brandes,  # Récupérer directement la valeur de la commune
                'repris_concurrence': commune.repris_concurrence,  # Récupérer directement la valeur de la commune
                'rapidite_deploiement': commune.get_rapidite_deploiement_value()
            }

        # Calcul des votes pour chaque commune
        for vote in commune.votes.all().select_related('user'):
            commune_results[commune.name]['qualite_site'] += vote.qualite_site
            commune_results[commune.name]['originalite_support'] += vote.originalite_support
            # Les critères administratifs sont déjà pris en compte, donc pas besoin de les additionner ici

    # Ajout des données dans la feuille Excel pour les communes
    for commune_name, results in commune_results.items():
        total_notes = (
            results['qualite_site'] +
            results['originalite_support'] +
            results['site_brandes'] +
            results['repris_concurrence'] +
            results['rapidite_deploiement']
        )
        row = [
            commune_name,
            results['qualite_site'],
            results['originalite_support'],
            results['site_brandes'],
            results['repris_concurrence'],
            results['rapidite_deploiement'],
            total_notes,
        ]
        ws_commune.append(row)

    # Feuille des résultats des sous-zones
    ws_sous_zone = wb.create_sheet(title="Résultats des Sous-Zones")
    headers_sous_zone = ['Sous-zone', 'Qualité du Site', 'Originalité du Support', 'site_brandes', 'repris_concurrence',
                         'rapidite_deploiement', 'Total']
    ws_sous_zone.append(headers_sous_zone)
    apply_header_format(ws_sous_zone)

    # Calcul des résultats pour chaque sous-zone
    for souszone in SousZone.objects.all():
        total_qualite_site = total_originalite_support = total_site_brandes = total_repris_concurrence = total_rapidite_deploiement = 0

        for commune in souszone.communes.all():
            for vote in commune.votes.all().select_related('user'):
                total_qualite_site += vote.qualite_site
                total_originalite_support += vote.originalite_support
                total_site_brandes += commune.site_brandes  # Récupérer directement la valeur de la commune
                total_repris_concurrence += commune.repris_concurrence  # Récupérer directement la valeur de la commune
                total_rapidite_deploiement += commune.get_rapidite_deploiement_value()  # Récupérer directement la valeur de la commune

        total_notes = total_qualite_site + total_originalite_support + total_site_brandes + total_repris_concurrence + total_rapidite_deploiement
        row = [souszone.name, total_qualite_site, total_originalite_support, total_site_brandes,
               total_repris_concurrence, total_rapidite_deploiement, total_notes]
        ws_sous_zone.append(row)

    # Feuille des résultats des zones
    ws_zone = wb.create_sheet(title="Résultats des Zones")
    headers_zone = ['Zone', 'Qualité du Site', 'Originalité du Support', 'site_brandes', 'repris_concurrence',
                    'rapidite_deploiement', 'Total']
    ws_zone.append(headers_zone)
    apply_header_format(ws_zone)

    # Calcul des résultats pour chaque zone
    for zone in Zone.objects.all():
        total_qualite_site = total_originalite_support = total_site_brandes = total_repris_concurrence = total_rapidite_deploiement = 0

        for souszone in zone.souszones.all():
            for commune in souszone.communes.all():
                for vote in commune.votes.all().select_related('user'):
                    total_qualite_site += vote.qualite_site
                    total_originalite_support += vote.originalite_support
                    total_site_brandes += commune.site_brandes  # Récupérer directement la valeur de la commune
                    total_repris_concurrence += commune.repris_concurrence  # Récupérer directement la valeur de la commune
                    total_rapidite_deploiement += commune.get_rapidite_deploiement_value()  # Récupérer directement la valeur de la commune

        total_notes = total_qualite_site + total_originalite_support + total_site_brandes + total_repris_concurrence + total_rapidite_deploiement
        row = [zone.nom, total_qualite_site, total_originalite_support, total_site_brandes, total_repris_concurrence,
               total_rapidite_deploiement, total_notes]
        ws_zone.append(row)

    # Envoi du fichier Excel dans la réponse HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="resultats_votes.xlsx"'

    wb.save(response)
    return response


class CommuneAdmin(admin.ModelAdmin):
    list_display = ('name', 'total_votes', 'note_moyenne')
    actions = [export_vote_results]

    def total_votes(self, obj):
        return obj.votes.count()

    total_votes.admin_order_field = 'votes__count'
    total_votes.short_description = 'Total des votes'

    def note_moyenne(self, obj):
        total_votes = obj.votes.count()
        if total_votes > 0:
            total_notes = sum(
                vote.qualite_site + vote.originalite_support + commune.site_brandes + commune.repris_concurrence + commune.get_rapidite_deploiement_value()
                for vote in obj.votes.all()
            )
            # Normalized by the number of fields (5 in this case)
            return total_notes / (total_votes * 5)
        return 0

    note_moyenne.short_description = 'Note moyenne'


# Enregistrer les modèles dans l'admin
admin.site.register(Commune, CommuneAdmin)
admin.site.register(Zone)
admin.site.register(SousZone)
admin.site.register(Vote)
admin.site.register(Image)
